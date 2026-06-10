"""
╔══════════════════════════════════════════════════════════════╗
║        LUNCH LADY — PDF ORDER IMPORT TOOL  v2                ║
║                                                              ║
║  • Integrated with Control Center System Architecture        ║
║  • All orders saved to Lunch Lady Template.xlsx              ║
║  • One sheet per month  (e.g. "June 2026")                  ║
║  • One master sheet with every order ever imported           ║
╚══════════════════════════════════════════════════════════════╝

INSTALL REQUIREMENTS (one time):
  pip install pdfplumber openpyxl
"""

import os, re, sys, glob, shutil
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

import pdfplumber
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                             Border, Side, GradientFill)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
#  CONFIGURATION  ← Dynamically calculated for the new structure
# ──────────────────────────────────────────────────────────────
# This file sits in: D:\VBA\Catering System\Lunch Lady\PY Files\
SCRIPT_DIR  = Path(__file__).parent

# Step up one folder to: D:\VBA\Catering System\Lunch Lady\
ROOT_DIR    = SCRIPT_DIR.parent

# New System Paths mapped from your folder hierarchy layout
PDF_FOLDER  = ROOT_DIR / "PDF Orders"
DONE_FOLDER = SCRIPT_DIR / "PDF_Orders_Done"
EXCEL_FILE  = ROOT_DIR / "Lunch Lady Template.xlsx"

# Column layout for each monthly sheet
# Every row = one lunch day for one child
COLUMNS = [
    "Order #",
    "Order Date",
    "Parent Name",
    "Student Name",
    "Teacher",
    "School",
    "Allergies",
    "Service Date",
    "Weekday",
    "Entree",
    "Vegetable / Side",
    "Fruit / Side 2",
]

# Colors
COLOR_HEADER_BG  = "1F4E79"   # dark blue
COLOR_HEADER_FG  = "FFFFFF"   # white
COLOR_ALT_ROW    = "DDEEFF"   # light blue alternate rows
COLOR_MONTH_TAB  = {           # tab colors by month number
    1:  "4472C4", 2:  "ED7D31", 3:  "A9D18E",
    4:  "FF0000", 5:  "FFC000", 6:  "70AD47",
    7:  "5B9BD5", 8:  "FF7F7F", 9:  "7030A0",
    10: "00B0F0", 11: "FF6600", 12: "C00000",
}

MASTER_SHEET = "All Orders"


# ──────────────────────────────────────────────────────────────
#  KNOWN DISH NAME MAP
#  Maps any recognised stub (lowercased) → canonical dish name.
#  Used by clean_entree() as a fast-path lookup BEFORE regex.
#  Add new dishes here whenever the menu changes.
# ──────────────────────────────────────────────────────────────
KNOWN_DISHES = {
    "teriyaki chicken bowl":            "Teriyaki Chicken Bowl",
    "chicken taco bowl":                "Chicken Taco Bowl",
    "chicken ranch wrap":               "Chicken Ranch Wrap",
    "bbq chicken sandwich":             "BBQ Chicken Sandwich",
    "chicken nuggets":                  "Chicken Nuggets",
    "chicken alfredo pasta":            "Chicken Alfredo Pasta",
    "pasta with meat sauce":            "Pasta with Meat Sauce",
    "mini chicken corn dogs":           "Mini Chicken Corn Dogs",
    "protein box":                      "Protein Box",
    "turkey pinwheels":                 "Turkey Pinwheels",
    "black bean & rice bowl":           "Black Bean & Rice Bowl",
    "black bean and rice bowl":         "Black Bean & Rice Bowl",
    "cheese & black bean quesadilla":   "Cheese & Black Bean Quesadilla",
    "cheese and black bean quesadilla": "Cheese & Black Bean Quesadilla",
    "whole grain french toast sticks with vanilla greek yogurt":
                                        "Whole Grain French Toast sticks with Vanilla Greek Yogurt",
}


# ──────────────────────────────────────────────────────────────
#  STYLING HELPERS
# ──────────────────────────────────────────────────────────────

def header_style():
    return {
        "font":      Font(bold=True, color=COLOR_HEADER_FG, size=11),
        "fill":      PatternFill("solid", fgColor=COLOR_HEADER_BG),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border":    Border(
            bottom=Side(style="medium", color="FFFFFF"),
            right= Side(style="thin",   color="FFFFFF"),
        ),
    }

def apply_style(cell, styles: dict):
    for attr, val in styles.items():
        setattr(cell, attr, val)

def style_header_row(ws, row=1):
    hs = header_style()
    for col in range(1, len(COLUMNS) + 1):
        apply_style(ws.cell(row, col), hs)
    ws.row_dimensions[row].height = 28

def style_data_row(ws, row: int):
    fill = PatternFill("solid", fgColor=COLOR_ALT_ROW) if row % 2 == 0 else None
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row, col)
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(
            bottom=Side(style="thin", color="DDDDDD"),
            right= Side(style="thin", color="DDDDDD"),
        )

def set_col_widths(ws):
    widths = [10, 13, 18, 18, 18, 32, 12, 14, 11, 28, 26, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


# ──────────────────────────────────────────────────────────────
#  WORKBOOK SETUP
# ──────────────────────────────────────────────────────────────

def get_or_create_workbook() -> openpyxl.Workbook:
    """Load existing workbook or create a fresh one."""
    if EXCEL_FILE.exists():
        return load_workbook(str(EXCEL_FILE))

    print(f"  → Creating new Excel file: {EXCEL_FILE.name}")
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    return wb


def get_or_create_sheet(wb, sheet_name: str, month_num: int = 0) -> openpyxl.worksheet.worksheet.Worksheet:
    """Return existing sheet or create it with headers."""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    # Tab color
    if month_num and month_num in COLOR_MONTH_TAB:
        ws.sheet_properties.tabColor = COLOR_MONTH_TAB[month_num]
    else:
        ws.sheet_properties.tabColor = "1F4E79"

    # Header row
    for i, col_name in enumerate(COLUMNS, 1):
        ws.cell(1, i).value = col_name

    style_header_row(ws)
    set_col_widths(ws)
    print(f"  → Created sheet: '{sheet_name}'")
    return ws


def ensure_master_sheet(wb):
    """Make sure the All Orders master sheet exists."""
    return get_or_create_sheet(wb, MASTER_SHEET, month_num=0)


# ──────────────────────────────────────────────────────────────
#  UNIVERSAL WIPE + WRITE  (used for ALL sheets — monthly + master)
#
#  THE FIX for row-drift on both monthly and All Orders sheets:
#
#  Old approach:  delete_existing_rows() scanned and deleted row
#  by row, leaving ghost formatting that pushed max_row to 100+.
#  Then append_rows_to_sheet() used max_row+1 → data landed at
#  row 107 / row 422 instead of row 2.
#
#  New approach:  wipe_sheet_data() physically deletes every row
#  from 2 onward in one call → max_row resets to 1 (header only).
#  write_rows_from_row2() then hardcodes the start index to 2,
#  so data always lands exactly where it should.
#
#  Daily_Packing / Kitchen_Prep formulas are safe because they
#  use whole-column refs like 'June 2026'!A:A or 'All Orders'!H:H
#  — those refs survive physical row deletion without any #REF!.
# ──────────────────────────────────────────────────────────────

def wipe_sheet_data(ws):
    """
    Physically delete every data row (row 2 onward) from the sheet.
    After this call ws.max_row == 1 (header only), so max_row drift
    is impossible on the next write.
    """
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)


def write_rows_from_row2(ws, rows: list):
    """
    Write a list of row-dicts into ws starting HARD at row 2.
    Never consults ws.max_row — immune to drift.
    """
    for i, row_data in enumerate(rows):
        r = 2 + i
        ws.cell(r,  1).value = row_data["order_num"]
        ws.cell(r,  2).value = row_data["order_date"]
        ws.cell(r,  2).number_format = "MMM D, YYYY"
        ws.cell(r,  3).value = row_data["parent"]
        ws.cell(r,  4).value = row_data["student"]
        ws.cell(r,  5).value = row_data["teacher"]
        ws.cell(r,  6).value = row_data["school"]
        ws.cell(r,  7).value = row_data["allergy"]
        ws.cell(r,  8).value = row_data["service_date"]
        ws.cell(r,  8).number_format = "MMM D, YYYY"
        ws.cell(r,  9).value = row_data["weekday"]
        ws.cell(r, 10).value = row_data["entree"]
        ws.cell(r, 11).value = row_data["side1"]
        ws.cell(r, 12).value = row_data["side2"]
        style_data_row(ws, r)


# ──────────────────────────────────────────────────────────────
#  PDF PARSING
# ──────────────────────────────────────────────────────────────

def extract_pdf_text(pdf_path: str) -> str:
    """
    Extract text from PDF with improved spacing.
    x_tolerance=1 prevents pdfplumber from merging words that are
    slightly close together (which causes 'Whole grainbreaded' fusing).
    """
    noise = {"Product Quantity Price", "Quantity Price", "Product  Quantity  Price"}
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text(x_tolerance=1)
            if t:
                text += t + "\n"

    clean = []
    for line in text.split("\n"):
        line = line.strip()
        if not line or line in noise:
            return text  # Safe recovery fallback
        # Fix fused words: insert space before capital after lowercase
        line = re.sub(r'([a-z])([A-Z])', r'\1 \2', line)
        # Fix fused words: lowercase letter jammed against uppercase+word
        line = re.sub(r'([a-z]{3,})([A-Z][a-z])', r'\1 \2', line)
        clean.append(line)
    return "\n".join(clean)


def parse_order_header(text: str) -> dict:
    order_num  = re.search(r'New order[:\s#]+(\d+)', text, re.IGNORECASE)
    order_date = re.search(r'Order #\d+\s*\(([^)]+)\)', text)
    parent     = re.search(r'new order from (.+?):', text, re.IGNORECASE)
    return {
        "order_num":  order_num.group(1).strip()  if order_num  else "UNKNOWN",
        "order_date": order_date.group(1).strip() if order_date else "",
        "parent":     parent.group(1).strip()     if parent     else "Unknown",
    }


def clean_entree(value: str) -> str:
    """
    Return the full entree text exactly as it appears in the PDF.
    Only fixes PDF word-fusion artifacts (e.g. "grainbreaded" -> "grain breaded").
    Nothing else is stripped — dish name, description, and diet tags all kept.
    """
    value = re.sub(r'([a-z])([A-Z])', r'\1 \2', value)
    value = re.sub(r'([a-z]{3,})([A-Z][a-z])', r'\1 \2', value)
    return value.strip()


def clean_side(value: str) -> str:
    value = re.sub(r'\s*\([^)]*\)\s*$', '', value)
    return value.strip()


def parse_children(text: str) -> list:
    blocks = re.split(r"Child's Name:", text)[1:]
    children = []

    for block in blocks:
        lines = [l.strip() for l in block.split("\n") if l.strip()]
        if not lines:
            continue

        child = {
            "name":    lines[0].strip(),
            "teacher": "",
            "school":  "",
            "allergy": "No",
            "menu":    {},
            "month":   "",
            "year":    datetime.now().year,
        }

        # Parse metadata
        i = 0
        while i < len(lines):
            line = lines[i]
            if "Teacher's Name:" in line:
                child["teacher"] = line.split("Teacher's Name:", 1)[1].strip()
            elif "Does Your Child Have Any Allergy?" in line:
                val = line.split(":", 1)[1].strip() if ":" in line else "No"
                child["allergy"] = val if val.lower() != "no" else "No"
            elif "Select Your Child" in line and "School:" in line:
                school_part = line.split("School:", 1)[1].strip()
                if (i + 1 < len(lines) and not re.match(
                        r'(Kindly|June|January|February|March|April|May|July|August|'
                        r'September|October|November|December|\(|Does|Teacher|Select|\$)',
                        lines[i + 1])):
                    school_part += " " + lines[i + 1]
                    i += 1
                child["school"] = school_part.strip()
            i += 1

        # Detect month
        month_match = re.search(
            r':\s*(January|February|March|April|May|June|July|August|'
            r'September|October|November|December)\s+\d+\s*\(\+',
            block, re.IGNORECASE)
        if month_match:
            child["month"] = month_match.group(1).capitalize()

        # Parse daily menu — join continuation lines first
        month_re = (r'January|February|March|April|May|June|July|August|'
                    r'September|October|November|December')
        # Handle both '(June 18)' and typo '(June 18' missing closing paren
        menu_re = re.compile(
            rf'\(({month_re})\s+(\d+)\)?\s*(Entree|Vegetables?|Fruits?)[\s:]+(.+)',
            re.IGNORECASE)

        # Detects any new (Month N) tag — used to stop line joining
        new_item_re = re.compile(
            rf'\(({month_re})\s+\d+\)',
            re.IGNORECASE)

        joined = []
        for line in lines:
            if menu_re.match(line):
                joined.append(line)
            elif joined and not re.match(
                    r"(Child's|Teacher's|Does|Select|Kindly|×|\$|Subtotal|"
                    r"Sales|Total|Payment|Delivery|Billing|Congratulations)", line):
                if new_item_re.match(line):
                    joined.append(line)
                else:
                    joined[-1] += " " + line
            else:
                joined.append(line)

        for line in joined:
            m = menu_re.match(line)
            if not m:
                continue
            month_name = m.group(1).capitalize()
            day_num    = int(m.group(2))
            mtype      = m.group(3).lower()
            value      = m.group(4).strip()

            if not child["month"]:
                child["month"] = month_name

            try:
                month_num = datetime.strptime(month_name, "%B").month
                day_date  = date(child["year"], month_num, day_num)
            except ValueError:
                continue

            if day_date not in child["menu"]:
                child["menu"][day_date] = {"entree": "", "side1": "", "side2": ""}

            if "entree" in mtype:
                child["menu"][day_date]["entree"] = clean_entree(value)
            elif "veg" in mtype:
                child["menu"][day_date]["side1"]  = clean_side(value)
            elif "fruit" in mtype:
                child["menu"][day_date]["side2"]  = clean_side(value)

        if child["menu"]:
            children.append(child)

    return children


# ──────────────────────────────────────────────────────────────
#  BUILD ROW DICTS FROM PARSED DATA
# ──────────────────────────────────────────────────────────────

def build_rows(header: dict, child: dict) -> list:
    """Convert parsed child data into a list of row dicts."""
    try:
        order_date_obj = datetime.strptime(header["order_date"], "%B %d, %Y")
    except (ValueError, KeyError):
        order_date_obj = datetime.now()

    rows = []
    for day_date in sorted(child["menu"].keys()):
        meals = child["menu"][day_date]
        rows.append({
            "order_num":    header["order_num"],
            "order_date":   order_date_obj,
            "parent":       header["parent"],
            "student":      child["name"],
            "teacher":      child["teacher"],
            "school":       child["school"],
            "allergy":      child["allergy"] if child["allergy"].lower() != "no" else "",
            "service_date": datetime(day_date.year, day_date.month, day_date.day),
            "weekday":      day_date.strftime("%A"),
            "entree":       meals.get("entree", ""),
            "side1":        meals.get("side1",  ""),
            "side2":        meals.get("side2",  ""),
        })
    return rows


# ──────────────────────────────────────────────────────────────
#  PROCESS ONE PDF
# ──────────────────────────────────────────────────────────────

def process_pdf(pdf_path: str, wb, monthly_collector: dict, master_collector: list) -> bool:
    """
    Parse one PDF and collect its rows into:
      • monthly_collector  — dict keyed by sheet_name → list of row dicts
      • master_collector   — flat list of ALL row dicts (for All Orders)
    No writing to the workbook happens here; all writes happen in main()
    after every PDF has been parsed, so each sheet is wiped once and
    written once cleanly.
    """
    filename = os.path.basename(pdf_path)
    print(f"\n{'='*60}")
    print(f"  Processing: {filename}")
    print(f"{'='*60}")

    text     = extract_pdf_text(pdf_path)
    header   = parse_order_header(text)
    children = parse_children(text)

    print(f"  Order #:  {header['order_num']}")
    print(f"  Parent:   {header['parent']}")
    print(f"  Children: {len(children)}")

    if not children:
        print("  ⚠ No children/menu data found in this PDF. Skipping.")
        return False

    for child in children:
        print(f"\n  → {child['name']} | {child['teacher']} | "
              f"{child['month']} {child['year']} | {len(child['menu'])} days")

        if not child["menu"]:
            print("    ⚠ No menu data. Skipping.")
            continue

        rows = build_rows(header, child)

        # Determine which monthly sheet this child belongs to
        sheet_name = f"{child['month']} {child['year']}"
        month_num  = datetime.strptime(child["month"], "%B").month

        # Store month_num alongside so we can create the sheet later
        if sheet_name not in monthly_collector:
            monthly_collector[sheet_name] = {"month_num": month_num, "rows": []}
        monthly_collector[sheet_name]["rows"].extend(rows)

        # Also collect for master sheet
        master_collector.extend(rows)

    return True


# ──────────────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────────────

def main():
    print("\n" + "═" * 60)
    print("   LUNCH LADY — PDF ORDER IMPORT TOOL  v2")
    print("═" * 60)

    PDF_FOLDER.mkdir(exist_ok=True)
    DONE_FOLDER.mkdir(exist_ok=True)

    pdf_files = sorted(glob.glob(str(PDF_FOLDER / "*.pdf")))
    if not pdf_files:
        print(f"\n⚠ No PDFs found in:  {PDF_FOLDER}")
        print("  Drop your order PDFs there and run again.")
        return

    print(f"\n📂 Found {len(pdf_files)} PDF(s) to process")
    print(f"📊 Output Template File: {EXCEL_FILE.name}")

    wb = get_or_create_workbook()

    # ── STEP 1: Parse ALL PDFs, collect rows by month + master ──
    # No sheet writes happen yet — we gather everything first so
    # each sheet gets exactly ONE wipe + ONE clean write pass.
    monthly_collector = {}   # { "June 2026": { month_num: 6, rows: [...] } }
    master_collector  = []   # flat list of all rows

    success = 0
    for pdf_path in pdf_files:
        try:
            if process_pdf(pdf_path, wb, monthly_collector, master_collector):
                success += 1
        except Exception as e:
            print(f"\n  ❌ ERROR in {os.path.basename(pdf_path)}: {e}")
            import traceback; traceback.print_exc()

    # ── STEP 2: Write monthly sheets — wipe then hard-write from row 2 ──
    print(f"\n{'─'*60}")
    print("  Writing monthly sheets...")
    for sheet_name, data in monthly_collector.items():
        ws = get_or_create_sheet(wb, sheet_name, data["month_num"])

        # Wipe all existing data rows so max_row resets to 1
        wipe_sheet_data(ws)

        # Sort by service date then student name for a clean layout
        sorted_rows = sorted(data["rows"], key=lambda r: (r["service_date"], r["student"]))

        # Write starting hard at row 2 — no max_row drift possible
        write_rows_from_row2(ws, sorted_rows)
        print(f"  ✓ {sheet_name}: {len(sorted_rows)} rows written from row 2")

    # ── STEP 3: Rebuild All Orders master sheet ──
    print(f"\n{'─'*60}")
    print("  Rebuilding All Orders master sheet...")
    master_ws = ensure_master_sheet(wb)

    # Wipe all existing data rows
    wipe_sheet_data(master_ws)

    # Sort all rows by service date then student name
    master_collector.sort(key=lambda r: (r["service_date"], r["student"]))

    # Write starting hard at row 2
    write_rows_from_row2(master_ws, master_collector)
    print(f"  ✓ All Orders: {len(master_collector)} rows written from row 2")

    # ── STEP 4: Auto-filter on every sheet ──
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    print(f"\n{'='*60}")
    print("  Saving template modifications...")

    # Retry loop — if Excel has the file open, ask user to close it
    while True:
        try:
            wb.save(str(EXCEL_FILE))
            print(f"  ✅ Saved:  {EXCEL_FILE.name}")
            break
        except PermissionError:
            print()
            print("  ⚠️  CANNOT SAVE — Excel file is open!")
            print(f"     Please CLOSE  {EXCEL_FILE.name}  in Excel first.")
            print("     Then press Enter to try again...")
            input("     > ")
            print("  Retrying save...")

    # Move processed PDFs into the target archive folder
    for pdf_path in pdf_files:
        try:
            dest = DONE_FOLDER / Path(pdf_path).name
            if dest.exists():
                stem = Path(pdf_path).stem
                dest = DONE_FOLDER / f"{stem}_{datetime.now().strftime('%H%M%S')}.pdf"
            shutil.move(pdf_path, dest)
        except Exception as e:
            print(f"  ⚠ Could not move {Path(pdf_path).name}: {e}")

    print(f"\n✅ Done!  {success}/{len(pdf_files)} PDF(s) imported.")
    print(f"   Processed PDFs moved to → {DONE_FOLDER.name}/")
    print("═" * 60)


if __name__ == "__main__":
    main()